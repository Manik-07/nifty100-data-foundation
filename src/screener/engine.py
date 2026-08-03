import sqlite3
import yaml
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


DATABASE = "db/nifty100.db"
CONFIG = "config/screener_config.yaml"


# --------------------------------------------------
# Database connection
# --------------------------------------------------

def get_connection():
    return sqlite3.connect(DATABASE)


# --------------------------------------------------
# Load screener configuration
# --------------------------------------------------

def load_config():

    with open(CONFIG, "r") as file:
        return yaml.safe_load(file)


# --------------------------------------------------
# Load database tables
# --------------------------------------------------

def load_data():

    conn = get_connection()

    # ---------------------------------------------
    # Load main financial ratios
    # ---------------------------------------------
    ratios = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn
    )

    # ---------------------------------------------
    # Load company master
    # ---------------------------------------------
    companies = pd.read_sql(
        "SELECT * FROM companies",
        conn
    )

    # ---------------------------------------------
    # Load Profit & Loss data
    # ---------------------------------------------
    profit_loss = pd.read_sql(
        """
        SELECT
            company_id,
            year,
            sales,
            net_profit,
            dividend_payout
        FROM profitandloss
        """,
        conn
    )

    # ---------------------------------------------
    # Load valuation / market-cap data
    # ---------------------------------------------
    market_cap = pd.read_sql(
        """
        SELECT
            company_id,
            year,
            market_cap_crore,
            pe_ratio,
            pb_ratio,
            dividend_yield_pct
        FROM market_cap
        """,
        conn
    )
    sectors = pd.read_sql(
        """
        SELECT
            company_id,
            broad_sector,
            sub_sector,
            market_cap_category
        FROM sectors
        """,
        conn
    )

    conn.close()

    # =============================================
    # Calculate 3-Year Revenue CAGR
    # =============================================

    revenue_cagr_3yr = calculate_revenue_cagr_3yr(
        profit_loss
    )

    # =============================================
    # Create common numeric year for merging
    # =============================================

    ratios["merge_year"] = pd.to_numeric(
        ratios["year"]
        .astype(str)
        .str.extract(r"(\d{4})")[0],
        errors="coerce"
    )

    profit_loss["merge_year"] = pd.to_numeric(
        profit_loss["year"]
        .astype(str)
        .str.extract(r"(\d{4})")[0],
        errors="coerce"
    )

    market_cap["merge_year"] = pd.to_numeric(
        market_cap["year"]
        .astype(str)
        .str.extract(r"(\d{4})")[0],
        errors="coerce"
    )

    # ---------------------------------------------
    # Remove original year columns
    # ---------------------------------------------

    profit_loss = profit_loss.drop(
        columns=["year"]
    )

    market_cap = market_cap.drop(
        columns=["year"]
    )

    # =============================================
    # Merge P&L data
    # =============================================

    ratios = ratios.merge(
        profit_loss,
        on=["company_id", "merge_year"],
        how="left"
    )

    # =============================================
    # Merge valuation data
    # =============================================

    ratios = ratios.merge(
        market_cap,
        on=["company_id", "merge_year"],
        how="left"
    )

    # =============================================
    # Merge 3-Year Revenue CAGR
    # =============================================

    ratios = ratios.merge(
        revenue_cagr_3yr,
        on="company_id",
        how="left"
    )

    # =============================================
    # Merge sector information
    # =============================================

    ratios = ratios.merge(
        sectors,
        on="company_id",
        how="left"
    )

    # =============================================
    # Merge ROCE from company master
    # =============================================

    roce_data = companies[
        ["id", "roce_percentage"]
    ].copy()

    roce_data = roce_data.rename(
        columns={
            "id": "company_id",
            "roce_percentage": "return_on_capital_employed_pct"
        }
    )

    ratios = ratios.merge(
        roce_data,
        on="company_id",
        how="left"
    )

    # =============================================
    # Calculate CFO / PAT Ratio
    # =============================================

    ratios["cfo_pat_ratio"] = (
        ratios["cash_from_operations_cr"]
        / ratios["net_profit"].replace(0, pd.NA)
    )

    return ratios, companies

def add_fcf_cagr(df, periods=5):
    """Add 5-period FCF CAGR for each company using historical FCF rows."""
    result = df.copy()

    result["fcf_year_num"] = pd.to_numeric(
        result["year"]
        .astype(str)
        .str.extract(r"(\d{4})")[0],
        errors="coerce"
    )

    result = result.sort_values(
        ["company_id", "fcf_year_num"]
    )

    result["fcf_start"] = (
        result.groupby("company_id")["free_cash_flow_cr"]
        .shift(periods)
    )

    valid = (
        (result["fcf_start"] > 0)
        & (result["free_cash_flow_cr"] > 0)
    )

    result["fcf_cagr_5yr"] = pd.NA

    result.loc[valid, "fcf_cagr_5yr"] = (
        (
            result.loc[valid, "free_cash_flow_cr"]
            / result.loc[valid, "fcf_start"]
        ) ** (1 / periods) - 1
    ) * 100

    result["fcf_cagr_5yr"] = pd.to_numeric(
        result["fcf_cagr_5yr"],
        errors="coerce"
    )

    return result


def calculate_revenue_cagr_3yr(df):

    result = df.copy()

    # Extract numeric year
    result["year_num"] = pd.to_numeric(
        result["year"]
        .astype(str)
        .str.extract(r"(\d{4})")[0],
        errors="coerce"
    )

    # Ignore rows such as TTM
    result = result.dropna(
        subset=["year_num", "sales"]
    )

    result = result.sort_values(
        ["company_id", "year_num"]
    )

    cagr_records = []

    for company_id, group in result.groupby("company_id"):

        group = group.sort_values("year_num")

        latest_row = group.iloc[-1]
        latest_year = int(latest_row["year_num"])

        start_year = latest_year - 3

        start_data = group[
            group["year_num"] == start_year
        ]

        if start_data.empty:
            continue

        start_sales = start_data.iloc[-1]["sales"]
        end_sales = latest_row["sales"]

        if (
            pd.notna(start_sales)
            and pd.notna(end_sales)
            and start_sales > 0
            and end_sales > 0
        ):

            cagr = (
                (end_sales / start_sales) ** (1 / 3) - 1
            ) * 100

            cagr_records.append(
                {
                    "company_id": company_id,
                    "revenue_cagr_3yr": round(cagr, 2)
                }
            )

    return pd.DataFrame(cagr_records)

def add_debt_declining_flag(df):
    """
    Adds debt_to_equity_declining_yoy.

    True when the company's latest D/E is lower
    than its previous financial year's D/E.
    """

    data = df.copy()

    # Extract numeric year
    data["numeric_year"] = pd.to_numeric(
        data["year"]
        .astype(str)
        .str.extract(r"(\d{4})")[0],
        errors="coerce"
    )

    # Ignore rows such as TTM with no numeric year
    data = data[data["numeric_year"].notna()].copy()

    # Sort chronologically
    data = data.sort_values(
        ["company_id", "numeric_year"]
    )

    # Previous year's D/E
    data["previous_debt_to_equity"] = (
        data.groupby("company_id")["debt_to_equity"].shift(1)
    )

    # Is D/E declining?
    data["debt_to_equity_declining_yoy"] = (
        data["debt_to_equity"].notna()
        & data["previous_debt_to_equity"].notna()
        & (
            data["debt_to_equity"]
            < data["previous_debt_to_equity"]
        )
    )

    return data



# ==================================================
# Keep latest financial year for each company
# ==================================================

def get_latest_company_data(df):
    data = df.copy()

    data["latest_year_num"] = pd.to_numeric(
        data["year"].astype(str).str.extract(r"(\d{4})")[0],
        errors="coerce"
    )

    data = data[data["latest_year_num"].notna()].copy()

    data = data.sort_values(
        ["company_id", "latest_year_num"],
        ascending=[True, False]
    )

    data = data.drop_duplicates(
        subset=["company_id"],
        keep="first"
    )

    return data


# ==================================================
# P10/P90 winsorisation + 0-100 normalisation
# ==================================================

def normalize_metric(series, inverse=False):
    values = pd.to_numeric(series, errors="coerce")
    valid = values.dropna()

    if valid.empty:
        return pd.Series(0.0, index=values.index)

    p10 = valid.quantile(0.10)
    p90 = valid.quantile(0.90)

    if pd.isna(p10) or pd.isna(p90) or p90 == p10:
        score = pd.Series(50.0, index=values.index)
        score[values.isna()] = 0.0
        return score

    clipped = values.clip(lower=p10, upper=p90)

    score = ((clipped - p10) / (p90 - p10)) * 100

    if inverse:
        score = 100 - score

    return score.clip(0, 100).fillna(0.0)


# ==================================================
# Day 17 composite quality score
# ==================================================

def calculate_composite_score(df):
    data = df.copy()

    # Profitability — 35%
    data["roe_score"] = normalize_metric(
        data["return_on_equity_pct"]
    )

    data["roce_score"] = normalize_metric(
        data["return_on_capital_employed_pct"]
    )

    data["npm_score"] = normalize_metric(
        data["net_profit_margin_pct"]
    )

    # Cash Quality — 30%
    data["fcf_cagr_score"] = normalize_metric(
        data["fcf_cagr_5yr"]
    )

    data["cfo_pat_score"] = normalize_metric(
        data["cfo_pat_ratio"]
    )

    data["fcf_positive_score"] = (
        pd.to_numeric(
            data["free_cash_flow_cr"],
            errors="coerce"
        ).fillna(0) > 0
    ).astype(int) * 100

    # Growth — 20%
    data["revenue_growth_score"] = normalize_metric(
        data["revenue_cagr_5yr"]
    )

    data["pat_growth_score"] = normalize_metric(
        data["pat_cagr_5yr"]
    )

    # Leverage — 15%
    data["debt_score"] = normalize_metric(
        data["debt_to_equity"],
        inverse=True
    )

    # Debt-free companies always receive maximum ICR score.
    data["icr_score"] = normalize_metric(
        data["interest_coverage"]
    )

    debt_free = (
        pd.to_numeric(
            data["debt_to_equity"],
            errors="coerce"
        ) == 0
    )

    data.loc[debt_free, "icr_score"] = 100.0

    # Weighted raw score
    data["raw_composite_score"] = (
        data["roe_score"] * 0.15
        + data["roce_score"] * 0.10
        + data["npm_score"] * 0.10
        + data["fcf_cagr_score"] * 0.15
        + data["cfo_pat_score"] * 0.10
        + data["fcf_positive_score"] * 0.05
        + data["revenue_growth_score"] * 0.10
        + data["pat_growth_score"] * 0.10
        + data["debt_score"] * 0.10
        + data["icr_score"] * 0.05
    ).clip(0, 100)

    # Sector-relative composite score.
    # Each company's raw score is normalised against its broad-sector peers.
    data["composite_quality_score"] = (
        data.groupby(
            "broad_sector",
            dropna=False
        )["raw_composite_score"]
        .transform(normalize_metric)
        .round(2)
    )

    return data


# ==================================================
# Day 17 export columns
# ==================================================

def get_export_columns(df):
    identifier_columns = [
        "company_id",
        "year",
        "broad_sector",
        "sub_sector"
    ]

    # 20 KPI columns required for the screener report.
    kpi_columns = [
        "sales",
        "net_profit",
        "market_cap_crore",
        "return_on_equity_pct",
        "return_on_capital_employed_pct",
        "net_profit_margin_pct",
        "operating_profit_margin_pct",
        "debt_to_equity",
        "interest_coverage",
        "asset_turnover",
        "free_cash_flow_cr",
        "fcf_cagr_5yr",
        "cfo_pat_ratio",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "eps_cagr_5yr",
        "pe_ratio",
        "pb_ratio",
        "dividend_yield_pct",
        "composite_quality_score"
    ]

    wanted = identifier_columns + kpi_columns
    return [column for column in wanted if column in df.columns]



# --------------------------------------------------
# Apply screener filters
# --------------------------------------------------

def apply_filters(df, preset):

    result = df.copy()

    # --------------------------------------------------
    # ROE
    # --------------------------------------------------
    if "roe_min" in preset:
        result = result[
            result["return_on_equity_pct"] >= preset["roe_min"]
        ]

        # --------------------------------------------------
    # Debt to Equity
    # Financial companies are normally exempt because
    # leverage works differently for banks/NBFCs.
    #
    # Exception:
    # debt_to_equity_max = 0 means genuinely debt-free,
    # so NO sector exemption is allowed.
    # --------------------------------------------------

    if "debt_to_equity_max" in preset:

        debt_limit = preset["debt_to_equity_max"]

        # Debt-Free screen: everybody must actually have D/E = 0
        if debt_limit == 0:

            result = result[
                result["debt_to_equity"].fillna(float("inf")) <= 0
            ]

        # Normal D/E screens: Financials are exempt
        else:

            is_financial = (
                result["broad_sector"]
                .fillna("")
                .str.strip()
                .str.lower()
                == "financials"
            )

            passes_debt_filter = (
                result["debt_to_equity"].notna()
                & (result["debt_to_equity"] <= debt_limit)
            )

            result = result[
                is_financial | passes_debt_filter
            ]

    # --------------------------------------------------
    # Revenue CAGR 5Y
    # --------------------------------------------------
    if "revenue_cagr_5yr_min" in preset:
        result = result[
            result["revenue_cagr_5yr"] >= preset["revenue_cagr_5yr_min"]
        ]

    # --------------------------------------------------
    # PAT CAGR 5Y
    # --------------------------------------------------
    if "pat_cagr_5yr_min" in preset:
        result = result[
            result["pat_cagr_5yr"] >= preset["pat_cagr_5yr_min"]
        ]
        # --------------------------------------------------
    # Operating Profit Margin (OPM)
    # --------------------------------------------------
    if "opm_min" in preset:
        result = result[
            result["operating_profit_margin_pct"]
            >= preset["opm_min"]
        ]

    # --------------------------------------------------
    # Interest Coverage Ratio (ICR)
    # --------------------------------------------------
    if "icr_min" in preset:

        # Debt-free companies automatically pass ICR
        icr_values = result["interest_coverage"].copy()

        debt_free = (
    result["debt_to_equity"].notna()
    & (result["debt_to_equity"] == 0)
)

        icr_values = icr_values.where(
            ~debt_free,
            float("inf")
        )

        result = result[
            icr_values >= preset["icr_min"]
        ]

    # --------------------------------------------------
    # EPS CAGR 5Y
    # --------------------------------------------------
    if "eps_cagr_5yr_min" in preset:
        result = result[
            result["eps_cagr_5yr"]
            >= preset["eps_cagr_5yr_min"]
        ]

    # --------------------------------------------------
    # Asset Turnover
    # --------------------------------------------------
    if "asset_turnover_min" in preset:
        result = result[
            result["asset_turnover"]
            >= preset["asset_turnover_min"]
        ]

    # --------------------------------------------------
    # Free Cash Flow
    # --------------------------------------------------
    if "free_cash_flow_min" in preset:
        result = result[
            result["free_cash_flow_cr"] >= preset["free_cash_flow_min"]
        ]

    # --------------------------------------------------
    # Sales
    # --------------------------------------------------
    if "sales_min" in preset:
        result = result[
            result["sales"] >= preset["sales_min"]
        ]

    # --------------------------------------------------
    # P/E Ratio
    # --------------------------------------------------
    if "pe_max" in preset:
        result = result[
            result["pe_ratio"] <= preset["pe_max"]
        ]

    # --------------------------------------------------
    # P/B Ratio
    # --------------------------------------------------
    if "pb_max" in preset:
        result = result[
            result["pb_ratio"] <= preset["pb_max"]
        ]

    # --------------------------------------------------
    # Dividend Yield
    # --------------------------------------------------
    if "dividend_yield_min" in preset:
        result = result[
            result["dividend_yield_pct"] >= preset["dividend_yield_min"]
        ]

    # --------------------------------------------------
    # Dividend Payout
    # --------------------------------------------------
    if "dividend_payout_min" in preset:
        result = result[
            result["dividend_payout"] >= preset["dividend_payout_min"]
        ]
    
    # Dividend Payout Maximum
    if "dividend_payout_max" in preset:
        result = result[
            result["dividend_payout"] <= preset["dividend_payout_max"]
    ]
    

    # --------------------------------------------------
    # Market Cap
    # --------------------------------------------------
    if "market_cap_min" in preset:
        result = result[
            result["market_cap_crore"] >= preset["market_cap_min"]
        ]

    # --------------------------------------------------
    # Net Profit
    # --------------------------------------------------
    if "net_profit_min" in preset:
        result = result[
            result["net_profit"] >= preset["net_profit_min"]
        ]
        
    # Revenue CAGR 3Y
    if "revenue_cagr_3yr_min" in preset:
       result = result[
          result["revenue_cagr_3yr"]
        >= preset["revenue_cagr_3yr_min"]
    ]
       
# --------------------------------------------------
# Debt-to-Equity Declining YoY
# Used by Turnaround Watch
# --------------------------------------------------
    if preset.get("debt_to_equity_declining_yoy", False):
       result = result[
         result["debt_to_equity_declining_yoy"] == True
    ]    
    
    
    
     # --------------------------------------------------
# Debt-to-Equity declining YoY
# --------------------------------------------------
    if preset.get("debt_to_equity_declining_yoy") is True:
       result = result[
         result["debt_to_equity_declining_yoy"] == True
    ]
    # --------------------------------------------------
    # Sort final result
    # --------------------------------------------------
    return result.sort_values(
        "composite_quality_score",
        ascending=False
    )
def format_screener_report(file_path, config):

    workbook = load_workbook(file_path)

    green_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE"
    )

    red_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE"
    )

    threshold_map = {
        "roe_min": ("return_on_equity_pct", "min"),
        "debt_to_equity_max": ("debt_to_equity", "max"),
        "free_cash_flow_min": ("free_cash_flow_cr", "min"),
        "revenue_cagr_5yr_min": ("revenue_cagr_5yr", "min"),
        "pat_cagr_5yr_min": ("pat_cagr_5yr", "min"),
        "opm_min": ("operating_profit_margin_pct", "min"),
        "pe_max": ("pe_ratio", "max"),
        "pb_max": ("pb_ratio", "max"),
        "dividend_yield_min": ("dividend_yield_pct", "min"),
        "icr_min": ("interest_coverage", "min"),
        "market_cap_min": ("market_cap_crore", "min"),
        "net_profit_min": ("net_profit", "min"),
        "eps_cagr_5yr_min": ("eps_cagr_5yr", "min"),
        "asset_turnover_min": ("asset_turnover", "min"),
        "sales_min": ("sales", "min"),
        "dividend_payout_max": ("dividend_payout", "max"),
        "revenue_cagr_3yr_min": ("revenue_cagr_3yr", "min")
    }

    for sheet in workbook.worksheets:

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        headers = {
            cell.value: cell.column
            for cell in sheet[1]
        }

        preset = config.get(sheet.title, {})

        for config_key, threshold in preset.items():

            if config_key not in threshold_map:
                continue

            column_name, direction = threshold_map[config_key]

            if column_name not in headers:
                continue

            column_number = headers[column_name]

            for row in range(2, sheet.max_row + 1):

                cell = sheet.cell(
                    row=row,
                    column=column_number
                )

                if cell.value is None:
                    continue

                try:
                    value = float(cell.value)
                    threshold_value = float(threshold)
                except (TypeError, ValueError):
                    continue

                if direction == "min":
                    passed = value >= threshold_value
                else:
                    passed = value <= threshold_value

                cell.fill = (
                    green_fill
                    if passed
                    else red_fill
                )

        for column_cells in sheet.columns:

            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            sheet.column_dimensions[column_letter].width = min(
                max_length + 2,
                30
            )

    workbook.save(file_path)

    print("Excel formatting completed.")


# ==================================================
# MAIN
# ==================================================

if __name__ == "__main__":

    print("=" * 60)
    print("Financial Screener Engine")
    print("=" * 60)

    # Load config
    config = load_config()

    # Load database
    ratios, companies = load_data()

    print("\nOriginal Financial Ratio Rows:", len(ratios))
    print("Companies:", ratios["company_id"].nunique())

    # --------------------------------------------------
    # Historical calculations must happen before
    # reducing the data to one latest row per company.
    # --------------------------------------------------

    ratios = add_fcf_cagr(ratios)
    ratios = add_debt_declining_flag(ratios)

    # Keep latest financial-year record per company
    ratios = get_latest_company_data(ratios)

    print("\nLatest Financial Ratio Rows:", len(ratios))
    print(
        "Latest Companies:",
        ratios["company_id"].nunique()
    )

    # --------------------------------------------------
    # Day 17 composite score
    # --------------------------------------------------

    ratios = calculate_composite_score(ratios)

    print(
        "\nComposite Score Range:",
        round(ratios["composite_quality_score"].min(), 2),
        "to",
        round(ratios["composite_quality_score"].max(), 2)
    )

    # --------------------------------------------------
    # Create output directory
    # --------------------------------------------------

    os.makedirs(
        "output",
        exist_ok=True
    )

    output_file = "output/screener_output.xlsx"

    # --------------------------------------------------
    # Run all six preset screeners
    # --------------------------------------------------

    print("\n" + "=" * 60)
    print("RUNNING ALL PRESET SCREENERS")
    print("=" * 60)

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        for preset_name, preset in config.items():

            result = apply_filters(
                ratios,
                preset
            )

            print(f"\n{preset_name.upper()}")
            print("-" * 50)
            print("Companies Found:", len(result))

            export_columns = get_export_columns(result)

            export_result = result[
                export_columns
            ].copy()

            export_result.to_excel(
                writer,
                sheet_name=preset_name[:31],
                index=False
            )

            if not result.empty:

                display_columns = [
                    "company_id",
                    "year",
                    "broad_sector",
                    "return_on_equity_pct",
                    "debt_to_equity",
                    "composite_quality_score"
                ]

                available_columns = [
                    column
                    for column in display_columns
                    if column in result.columns
                ]

                print(
                    result[
                        available_columns
                    ].head(10).to_string(index=False)
                )

            else:
                print("No companies matched this preset.")

    print("\n" + "=" * 60)
    print(f"Excel report saved to: {output_file}")
    print("=" * 60)

    format_screener_report(
        output_file,
        config
    )
