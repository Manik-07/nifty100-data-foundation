import os
import re
import sqlite3

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ==========================================================
# CONFIGURATION
# ==========================================================

DB_PATH = "db/nifty100.db"
PEER_GROUP_PATH = "data/raw/peer_groups.xlsx"
OUTPUT_PATH = "output/peer_comparison.xlsx"


# ==========================================================
# DATABASE CONNECTION
# ==========================================================

def get_connection():
    return sqlite3.connect(DB_PATH)


# ==========================================================
# LOAD DATA
# ==========================================================

def load_data():

    conn = get_connection()

    ratios = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn
    )

    companies = pd.read_sql(
        "SELECT * FROM companies",
        conn
    )

    conn.close()

    peer_groups = pd.read_excel(
        PEER_GROUP_PATH
    )

    return ratios, companies, peer_groups


# ==========================================================
# EXTRACT YEAR
# ==========================================================

def extract_year(value):

    match = re.search(
        r"\d{4}",
        str(value)
    )

    if match:
        return int(match.group())

    return None


# ==========================================================
# GET LATEST FINANCIAL RATIOS
# ==========================================================

def get_latest_ratios(ratios):

    df = ratios.copy()

    df["year_num"] = (
        df["year"]
        .apply(extract_year)
    )

    df = df[
        df["year_num"].notna()
    ].copy()

    df = (
        df
        .sort_values(
            [
                "company_id",
                "year_num"
            ]
        )
        .groupby(
            "company_id",
            as_index=False
        )
        .tail(1)
        .reset_index(drop=True)
    )

    return df


# ==========================================================
# PREPARE PEER DATA
# ==========================================================

def prepare_peer_data(
    latest,
    companies,
    peer_groups
):

    latest = latest.copy()
    companies = companies.copy()
    peer_groups = peer_groups.copy()

    # ------------------------------------------------------
    # Clean company IDs
    # ------------------------------------------------------

    latest["company_id"] = (
        latest["company_id"]
        .astype(str)
        .str.strip()
    )

    companies["id"] = (
        companies["id"]
        .astype(str)
        .str.strip()
    )

    peer_groups["company_id"] = (
        peer_groups["company_id"]
        .astype(str)
        .str.strip()
    )

    peer_groups["peer_group_name"] = (
        peer_groups["peer_group_name"]
        .astype(str)
        .str.strip()
    )

    peer_groups["is_benchmark"] = (
        pd.to_numeric(
            peer_groups["is_benchmark"],
            errors="coerce"
        )
        .fillna(0)
        .astype(int)
    )

    # ------------------------------------------------------
    # Merge peer-group mapping
    # ------------------------------------------------------

    df = latest.merge(
        peer_groups[
            [
                "company_id",
                "peer_group_name",
                "is_benchmark"
            ]
        ],
        on="company_id",
        how="inner"
    )

    # ------------------------------------------------------
    # Company name + ROCE
    # ------------------------------------------------------

    company_columns = [
        "id",
        "company_name",
        "roce_percentage"
    ]

    company_info = (
        companies[
            company_columns
        ]
        .copy()
        .rename(
            columns={
                "id": "company_id",
                "roce_percentage":
                    "return_on_capital_employed_pct"
            }
        )
    )

    df = df.merge(
        company_info,
        on="company_id",
        how="left"
    )

    return df


# ==========================================================
# ADD PEER PERCENTILE RANKS
# ==========================================================

def add_percentile_ranks(df):

    df = df.copy()

    # Day 18 / Day 20 required peer metrics
    metrics = [
        "return_on_equity_pct",
        "return_on_capital_employed_pct",
        "net_profit_margin_pct",
        "debt_to_equity",
        "free_cash_flow_cr",
        "pat_cagr_5yr",
        "revenue_cagr_5yr",
        "eps_cagr_5yr",
        "interest_coverage",
        "asset_turnover"
    ]

    for metric in metrics:

        if metric not in df.columns:

            print(
                f"WARNING: Missing metric: {metric}"
            )

            continue

        # Convert to numeric
        df[metric] = pd.to_numeric(
            df[metric],
            errors="coerce"
        )

        percentile_column = (
            f"{metric}_percentile"
        )

        # --------------------------------------------------
        # D/E inverse percentile
        # Lower D/E = better
        # --------------------------------------------------

        if metric == "debt_to_equity":

            df[percentile_column] = (
                df.groupby(
                    "peer_group_name"
                )[metric]
                .rank(
                    pct=True,
                    ascending=False
                )
                * 100
            )

        # --------------------------------------------------
        # Higher is better
        # --------------------------------------------------

        else:

            df[percentile_column] = (
                df.groupby(
                    "peer_group_name"
                )[metric]
                .rank(
                    pct=True,
                    ascending=True
                )
                * 100
            )

        df[percentile_column] = (
            df[percentile_column]
            .round(2)
        )

    return df


# ==========================================================
# REPORT COLUMNS
# ==========================================================

def get_report_columns(df):

    # Core identification columns
    columns = [
        "company_id",
        "company_name",
        "peer_group_name",
        "is_benchmark",
        "year"
    ]

    # ------------------------------------------------------
    # 20 financial/KPI columns
    # ------------------------------------------------------

    metric_columns = [
        "return_on_equity_pct",
        "return_on_capital_employed_pct",
        "net_profit_margin_pct",
        "operating_profit_margin_pct",
        "debt_to_equity",
        "interest_coverage",
        "asset_turnover",
        "free_cash_flow_cr",
        "cash_from_operations_cr",
        "capex_cr",
        "pat_cagr_5yr",
        "revenue_cagr_5yr",
        "eps_cagr_5yr",
        "earnings_per_share",
        "book_value_per_share",
        "dividend_payout_ratio_pct",
        "total_debt_cr",
        "composite_quality_score"
    ]

    # ------------------------------------------------------
    # 10 percentile columns
    # ------------------------------------------------------

    percentile_columns = [
        "return_on_equity_pct_percentile",
        "return_on_capital_employed_pct_percentile",
        "net_profit_margin_pct_percentile",
        "debt_to_equity_percentile",
        "free_cash_flow_cr_percentile",
        "pat_cagr_5yr_percentile",
        "revenue_cagr_5yr_percentile",
        "eps_cagr_5yr_percentile",
        "interest_coverage_percentile",
        "asset_turnover_percentile"
    ]

    for column in metric_columns:

        if (
            column in df.columns
            and column not in columns
        ):
            columns.append(column)

    for column in percentile_columns:

        if (
            column in df.columns
            and column not in columns
        ):
            columns.append(column)

    return columns


# ==========================================================
# EXPORT PEER COMPARISON EXCEL
# ==========================================================

def export_peer_comparison(df):

    os.makedirs(
        "output",
        exist_ok=True
    )

    report_columns = (
        get_report_columns(df)
    )

    with pd.ExcelWriter(
        OUTPUT_PATH,
        engine="openpyxl"
    ) as writer:

        for peer_group in sorted(
            df[
                "peer_group_name"
            ].dropna().unique()
        ):

            group_df = df[
                df[
                    "peer_group_name"
                ] == peer_group
            ].copy()

            # --------------------------------------------------
            # Sort benchmark first, then composite score
            # --------------------------------------------------

            sort_columns = []
            ascending = []

            if "is_benchmark" in group_df.columns:

                sort_columns.append(
                    "is_benchmark"
                )

                ascending.append(
                    False
                )

            if (
                "composite_quality_score"
                in group_df.columns
            ):

                sort_columns.append(
                    "composite_quality_score"
                )

                ascending.append(
                    False
                )

            if sort_columns:

                group_df = (
                    group_df
                    .sort_values(
                        sort_columns,
                        ascending=ascending
                    )
                )

            # --------------------------------------------------
            # Select report columns
            # --------------------------------------------------

            group_df = group_df[
                [
                    column
                    for column in report_columns
                    if column in group_df.columns
                ]
            ]

            # Excel sheet names max 31 chars
            sheet_name = (
                str(peer_group)[:31]
            )

            group_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

            print(
                f"{peer_group}: "
                f"{len(group_df)} companies"
            )

    print(
        "\nExcel report generated:",
        OUTPUT_PATH
    )


# ==========================================================
# FORMAT EXCEL REPORT
# ==========================================================

def format_peer_report():

    workbook = load_workbook(
        OUTPUT_PATH
    )

    # ------------------------------------------------------
    # FILLS
    # ------------------------------------------------------

    green_fill = PatternFill(
        "solid",
        fgColor="C6EFCE"
    )

    yellow_fill = PatternFill(
        "solid",
        fgColor="FFEB9C"
    )

    red_fill = PatternFill(
        "solid",
        fgColor="FFC7CE"
    )

    gold_fill = PatternFill(
        "solid",
        fgColor="FFD966"
    )

    header_fill = PatternFill(
        "solid",
        fgColor="D9EAF7"
    )

    median_fill = PatternFill(
        "solid",
        fgColor="D9EAD3"
    )

    # ------------------------------------------------------
    # FORMAT EACH PEER SHEET
    # ------------------------------------------------------

    for sheet in workbook.worksheets:

        # --------------------------------------------------
        # HEADER
        # --------------------------------------------------

        for cell in sheet[1]:

            cell.font = Font(
                bold=True
            )

            cell.fill = (
                header_fill
            )

            cell.alignment = (
                Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
            )

        headers = {
            cell.value: cell.column
            for cell in sheet[1]
        }

        percentile_columns = [
            name
            for name in headers
            if str(name).endswith(
                "_percentile"
            )
        ]

        benchmark_column = (
            headers.get(
                "is_benchmark"
            )
        )

        # --------------------------------------------------
        # PERCENTILE COLOURS
        #
        # >=75 = green
        # 25-75 = yellow
        # <=25 = red
        # --------------------------------------------------

        for row in range(
            2,
            sheet.max_row + 1
        ):

            for column_name in (
                percentile_columns
            ):

                column = (
                    headers[
                        column_name
                    ]
                )

                cell = sheet.cell(
                    row=row,
                    column=column
                )

                value = cell.value

                if value is None:
                    continue

                try:
                    value = float(value)

                except (
                    TypeError,
                    ValueError
                ):
                    continue

                if value >= 75:

                    cell.fill = (
                        green_fill
                    )

                elif value <= 25:

                    cell.fill = (
                        red_fill
                    )

                else:

                    cell.fill = (
                        yellow_fill
                    )

        # --------------------------------------------------
        # BENCHMARK ROW
        # --------------------------------------------------

        if benchmark_column:

            for row in range(
                2,
                sheet.max_row + 1
            ):

                benchmark = (
                    sheet.cell(
                        row=row,
                        column=benchmark_column
                    ).value
                )

                if benchmark in (
                    True,
                    1,
                    "1",
                    "TRUE",
                    "True"
                ):

                    for cell in (
                        sheet[row]
                    ):

                        cell.fill = (
                            gold_fill
                        )

                        cell.font = Font(
                            bold=True
                        )

        # --------------------------------------------------
        # MEDIAN SUMMARY ROW
        # --------------------------------------------------

        median_row = (
            sheet.max_row + 2
        )

        sheet.cell(
            row=median_row,
            column=1,
            value="PEER MEDIAN"
        )

        sheet.cell(
            row=median_row,
            column=1
        ).font = Font(
            bold=True
        )

        # --------------------------------------------------
        # Calculate median for numeric columns
        # --------------------------------------------------

        for column in range(
            1,
            sheet.max_column + 1
        ):

            values = []

            for row in range(
                2,
                median_row - 1
            ):

                value = (
                    sheet.cell(
                        row=row,
                        column=column
                    ).value
                )

                if (
                    isinstance(
                        value,
                        (int, float)
                    )
                    and not isinstance(
                        value,
                        bool
                    )
                ):

                    values.append(
                        value
                    )

            if values:

                series = pd.Series(
                    values,
                    dtype=float
                )

                median = (
                    series.median()
                )

                sheet.cell(
                    row=median_row,
                    column=column,
                    value=round(
                        float(median),
                        2
                    )
                )

        # --------------------------------------------------
        # Median styling
        # --------------------------------------------------

        for cell in (
            sheet[median_row]
        ):

            cell.fill = (
                median_fill
            )

            cell.font = Font(
                bold=True
            )

        # --------------------------------------------------
        # FREEZE HEADER
        # --------------------------------------------------

        sheet.freeze_panes = (
            "A2"
        )

        # --------------------------------------------------
        # AUTO FILTER
        # --------------------------------------------------

        last_data_row = (
            median_row - 2
        )

        sheet.auto_filter.ref = (
            f"A1:"
            f"{get_column_letter(sheet.max_column)}"
            f"{last_data_row}"
        )

        # --------------------------------------------------
        # COLUMN WIDTHS
        # --------------------------------------------------

        for column_cells in (
            sheet.columns
        ):

            max_length = 0

            column_letter = (
                get_column_letter(
                    column_cells[
                        0
                    ].column
                )
            )

            for cell in (
                column_cells
            ):

                if (
                    cell.value
                    is not None
                ):

                    max_length = max(
                        max_length,
                        len(
                            str(
                                cell.value
                            )
                        )
                    )

            sheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                30
            )

        # Header row height
        sheet.row_dimensions[
            1
        ].height = 35

    # ------------------------------------------------------
    # SAVE
    # ------------------------------------------------------

    workbook.save(
        OUTPUT_PATH
    )

    print(
        "\nExcel formatting completed."
    )


# ==========================================================
# VALIDATE GENERATED WORKBOOK
# ==========================================================

def validate_workbook():

    workbook = load_workbook(
        OUTPUT_PATH,
        read_only=True
    )

    sheets = (
        workbook.sheetnames
    )

    print(
        "\n" + "=" * 60
    )

    print(
        "DAY 20 VALIDATION"
    )

    print(
        "=" * 60
    )

    print(
        "Excel sheets:",
        len(sheets)
    )

    print(
        "Sheet names:"
    )

    for sheet in sheets:
        print(" -", sheet)

    workbook.close()

    if len(sheets) == 11:

        print(
            "\n11 peer-group sheets: PASS"
        )

    else:

        print(
            "\n11 peer-group sheets: FAIL"
        )


# ==========================================================
# MAIN
# ==========================================================

if __name__ == "__main__":

    print("=" * 60)
    print(
        "DAY 20 - PEER COMPARISON EXCEL REPORT"
    )
    print("=" * 60)

    # ------------------------------------------------------
    # 1. LOAD DATA
    # ------------------------------------------------------

    (
        ratios,
        companies,
        peer_groups
    ) = load_data()

    print(
        "\nFinancial ratio rows:",
        len(ratios)
    )

    print(
        "Companies:",
        ratios[
            "company_id"
        ].nunique()
    )

    print(
        "\nPeer group rows:",
        len(peer_groups)
    )

    print(
        "Peer groups:",
        peer_groups[
            "peer_group_name"
        ].nunique()
    )

    # ------------------------------------------------------
    # 2. LATEST RATIOS
    # ------------------------------------------------------

    latest = (
        get_latest_ratios(
            ratios
        )
    )

    print(
        "\nLatest company records:",
        len(latest)
    )

    # ------------------------------------------------------
    # 3. PREPARE PEER DATA
    # ------------------------------------------------------

    peer_data = (
        prepare_peer_data(
            latest,
            companies,
            peer_groups
        )
    )

    print(
        "\nCompanies mapped to peer groups:",
        len(peer_data)
    )

    print(
        "Peer groups in final data:",
        peer_data[
            "peer_group_name"
        ].nunique()
    )

    # ------------------------------------------------------
    # ROCE VALIDATION
    # ------------------------------------------------------

    print(
        "Companies with ROCE:",
        peer_data[
            "return_on_capital_employed_pct"
        ].notna().sum()
    )

    missing_roce = (
        peer_data[
            peer_data[
                "return_on_capital_employed_pct"
            ].isna()
        ][
            ["company_id"]
        ]
    )

    if not missing_roce.empty:

        print(
            "\nPeer companies missing ROCE:"
        )

        print(
            missing_roce.to_string(
                index=False
            )
        )

    # ------------------------------------------------------
    # 4. CALCULATE PERCENTILES
    # ------------------------------------------------------

    peer_data = (
        add_percentile_ranks(
            peer_data
        )
    )

    percentile_columns = [
        column
        for column
        in peer_data.columns
        if column.endswith(
            "_percentile"
        )
    ]

    print(
        "\nPercentile columns created:",
        len(
            percentile_columns
        )
    )

    print(
        percentile_columns
    )

    # ------------------------------------------------------
    # 5. PEER GROUP COUNTS
    # ------------------------------------------------------

    print(
        "\nPeer group company counts:"
    )

    counts = (
        peer_data
        .groupby(
            "peer_group_name"
        )[
            "company_id"
        ]
        .nunique()
        .sort_index()
    )

    print(
        counts.to_string()
    )

    # ------------------------------------------------------
    # 6. EXPORT
    # ------------------------------------------------------

    export_peer_comparison(
        peer_data
    )

    # ------------------------------------------------------
    # 7. FORMAT
    # ------------------------------------------------------

    format_peer_report()

    # ------------------------------------------------------
    # 8. VALIDATE
    # ------------------------------------------------------

    validate_workbook()

    print(
        "\n" + "=" * 60
    )

    print(
        "DAY 20 REPORT GENERATION COMPLETED"
    )

    print(
        "=" * 60
    )